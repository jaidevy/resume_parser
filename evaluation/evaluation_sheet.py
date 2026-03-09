"""
Evaluation Sheet Generator.

Creates a comprehensive comparison of extracted vs expected fields
for all sample resumes. Supports JSON, CSV, and Excel (.xlsx) output.

This module fulfills Section 10 of the assignment:
"Show a comparison table of extracted vs expected values for 5+ resumes."
"""
import csv
import json
import os
import re
import sys
from datetime import datetime
from typing import Optional

# Force UTF-8 output on Windows (avoids cp1252 emoji crash)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]

# Add project root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "resume_parser.settings")

import django
django.setup()

from api.services.parser import ResumeParser
from api.services.extractor import ResumeExtractor


# ─── Expected Values for Each Sample Resume ───────────────────────────────────

EXPECTED_DATA = {
    "resume_01_standard.txt": {
        "candidate_full_name": "John Michael Doe",
        "email_ids": ["john.doe@email.com"],
        "phones": ["+1-555-123-4567"],
        "current_location": "San Francisco, CA, USA",
        "linkedin_url": "https://linkedin.com/in/johndoe",
        "github_url": "https://github.com/johndoe",
        "total_experience_years": 7,
        "key_skills": ["Python", "Django", "Flask", "JavaScript", "React", "TypeScript",
                       "PostgreSQL", "MongoDB", "Redis", "Docker", "Kubernetes", "AWS",
                       "Azure", "CI/CD", "Git", "REST APIs", "GraphQL", "Agile", "Scrum", "TDD"],
        "work_experience_count": 3,
        "education_count": 1,
        "certifications": ["AWS Certified Solutions Architect - Associate",
                          "Microsoft Certified: Azure Developer Associate",
                          "Certified Kubernetes Administrator (CKA)"],
        "languages_known": ["English", "Spanish", "French"],
        "gender": None,
    },
    "resume_02_data_analyst.txt": {
        "candidate_full_name": "Sarah Johnson",
        "email_ids": ["sarah.j@outlook.com", "sarah.johnson@company.com"],
        "phones": ["(212) 555-0198"],
        "current_location": "New York, NY",
        "linkedin_url": "https://linkedin.com/in/sarahjohnson-data",
        "github_url": "https://github.com/sjohnson-data",
        "total_experience_years": 4,
        "key_skills_contains": ["Power BI", "Power Automate", "Power Apps", "Python", "SQL", "R"],
        "work_experience_count": 3,
        "education_count": 2,
        "certifications_contains": ["PL-900", "DA-100"],
        "languages_known": ["English", "Mandarin"],
        "gender": None,
    },
    "resume_03_cloud_architect.txt": {
        "candidate_full_name": "Michael Chen",
        "email_ids": ["michael.chen@gmail.com"],
        "phones": ["+44-7700-900456"],
        "current_location": "London, United Kingdom",
        "total_experience_years": 10,
        "key_skills_contains": ["Azure", "AWS", "Kubernetes", "Docker", "Terraform", "Python", "Go"],
        "work_experience_count": 4,
        "education_count": 2,
        "certifications_contains": ["AWS Solutions Architect Professional", "CKA"],
        "languages_known": ["English", "Mandarin", "Cantonese"],
        "gender": None,
    },
    "resume_04_ux_designer.txt": {
        "candidate_full_name": "Maria Garcia Lopez",
        "email_ids": ["maria.garcia@email.com"],
        "phones": ["+34 612 345 678"],
        "current_location": "Barcelona, Spain",
        "total_experience_years": 3,
        "key_skills_contains": ["Figma", "React", "Vue.js", "JavaScript", "CSS3", "HTML5"],
        "work_experience_count": 2,
        "education_count": 1,
        "languages_known": ["Spanish", "Catalan", "English", "Portuguese"],
        "gender": None,
    },
    "resume_05_fullstack_dev.txt": {
        "candidate_full_name": "Rajesh Kumar Patel",
        "email_ids": ["rajesh.patel@email.com"],
        "phones": ["+91-9876543210"],
        "current_location": "Bangalore, Karnataka, India",
        "linkedin_url": "https://linkedin.com/in/rajeshpatel",
        "total_experience_years": 6,
        "key_skills_contains": ["Java", "Spring Boot", "Angular", "React", "Python", "SQL", "Docker"],
        "work_experience_count": 3,
        "education_count": 1,
        "certifications_contains": ["Oracle Certified Professional", "AWS Certified Developer"],
        "languages_known": ["English", "Hindi", "Kannada", "Gujarati"],
        "gender": "Male",
    },
    "resume_06_project_manager.txt": {
        "candidate_full_name": "Emily Watson",
        "email_ids": ["emily.watson92@gmail.com"],
        "phones": ["1-303-555-0147"],
        "current_location": "Denver, Colorado",
        "total_experience_years": 8,
        "key_skills_contains": ["Power Automate", "Power Apps", "Power BI", "Jira", "Azure DevOps"],
        "work_experience_count": 3,
        "education_count": 2,
        "certifications_contains": ["PMP", "CSM", "PL-900"],
        "languages_known": ["English"],
        "gender": None,
    },
    "resume_07_ml_engineer.txt": {
        "candidate_full_name": "Alex Kim",
        "email_ids": ["alex.kim@protonmail.com"],
        "phones": ["+82-10-1234-5678"],
        "current_location": "Seoul, South Korea",
        "linkedin_url": "https://linkedin.com/in/alexkim-ml",
        "github_url": "https://github.com/alexkim-ml",
        "total_experience_years": 5,
        "key_skills_contains": ["PyTorch", "TensorFlow", "Python", "Kubernetes", "Docker"],
        "work_experience_count": 3,
        "education_count": 2,
        "certifications_contains": ["AWS Machine Learning Specialty"],
        "languages_known": ["Korean", "English", "Japanese"],
        "gender": None,
    },
    "resume_08_fresh_graduate.txt": {
        "candidate_full_name": "Priya Sharma",
        "email_ids": ["priya.sharma24@gmail.com"],
        "phones": ["+91-8765432109"],
        "current_location": "New Delhi, India",
        "github_url_contains": "priyasharma",
        "total_experience_years": 0,
        "key_skills_contains": ["Python", "Java", "C++", "JavaScript", "React", "Django"],
        "work_experience_count": 3,
        "education_count": 1,
        "languages_known": ["English", "Hindi"],
        "gender": None,
    },
    "resume_09_devops_sre.txt": {
        "candidate_full_name": "James O'Brien",
        "email_ids": ["james.obrien@devops.io"],
        "phones": ["+1 (415) 555-2847"],
        "current_location": "San Jose, CA, USA",
        "linkedin_url": "https://linkedin.com/in/jamesobrien-sre",
        "github_url": "https://github.com/jobrien-devops",
        "total_experience_years": 9,
        "key_skills_contains": ["Terraform", "Kubernetes", "Docker", "AWS", "Python", "Go"],
        "work_experience_count": 4,
        "education_count": 1,
        "certifications_contains": ["CKA", "CKS", "AWS Solutions Architect Professional"],
        "languages_known": ["English", "Irish"],
        "gender": None,
    },
    "resume_10_cybersecurity.txt": {
        "candidate_full_name": "Aisha Okafor",  # LLM may include Dr./PhD prefix/suffix
        "email_ids": ["aisha.okafor@securitylab.org"],
        "phones": ["+234-802-123-4567"],
        "current_location": "Lagos, Nigeria",
        "linkedin_url": "https://linkedin.com/in/aishaokafor",
        "github_url": "https://github.com/aisha-security",
        "total_experience_years": 12,
        "key_skills_contains": ["Penetration Testing", "Python", "SIEM", "Burp Suite"],
        "work_experience_count": 4,
        "education_count": 3,
        "certifications_contains": ["OSCP", "CISSP", "CEH"],
        "languages_known": ["English", "Yoruba", "French", "Arabic"],
        "gender": None,
    },
    # ── Edge-case resumes ────────────────────────────────────────────────
    "resume_11_missing_phone.txt": {
        "candidate_full_name": "Daniela Rossi",
        "email_ids": ["daniela.rossi@email.com"],
        "phones": [],  # No phone number in resume
        "current_location": "Milan, Italy",
        "linkedin_url": "https://linkedin.com/in/danielarossi",
        "total_experience_years": 4,
        "key_skills_contains": ["Figma", "Sketch", "Adobe XD", "HTML5", "CSS3", "React"],
        "work_experience_count": 2,
        "education_count": 2,
        "certifications_contains": ["Google UX Design"],
        "gender": None,
    },
    "resume_12_multiple_emails.txt": {
        "candidate_full_name": "Ahmed Al-Rashidi",
        "email_ids": ["ahmed.rashidi@techcorp.com", "ahmed.r@gmail.com", "a.rashidi@alumni.kaust.edu.sa"],
        "phones": ["+966 55 123 4567", "+966 50 987 6543"],
        "current_location": "Riyadh, Saudi Arabia",
        "linkedin_url": "https://linkedin.com/in/ahmedrashidi",
        "github_url": "https://github.com/arashidi",
        "total_experience_years": 6,
        "key_skills_contains": ["Python", "SQL", "Spark", "Kafka", "Airflow", "AWS", "Azure"],
        "work_experience_count": 3,
        "education_count": 2,
        "certifications_contains": ["AWS Certified Data Analytics", "DP-203"],
        "gender": None,
    },
    "resume_13_table_format.txt": {
        "candidate_full_name": "Lisa Andersson",
        "email_ids": ["lisa.andersson@outlook.com"],
        "phones": ["+46 70 123 4567"],
        "current_location": "Stockholm, Sweden",
        "linkedin_url": "https://linkedin.com/in/lisaandersson",
        "total_experience_years": 9,
        "key_skills_contains": ["Agile", "Jira", "Confluence", "Power BI", "SAFe"],
        "work_experience_count": 3,
        "education_count": 2,
        "certifications_contains": ["PMP", "CSM", "SPC"],
        "gender": None,
    },
    "resume_14_two_column.txt": {
        "candidate_full_name": "Carlos Mendes",
        "email_ids": ["carlos.mendes@email.com"],
        "phones": ["+55 11 98765-4321"],
        "current_location": "São Paulo, SP, Brazil",
        "total_experience_years": 5,
        "key_skills_contains": ["JavaScript", "TypeScript", "Python", "React", "Node.js", "Docker"],
        "work_experience_count": 3,
        "education_count": 1,
        "certifications_contains": ["AWS Solutions Architect"],
        "gender": None,
    },
    "resume_15_multipage.txt": {
        "candidate_full_name": "Victoria Blackwood",  # LLM may include Ph.D. suffix
        "email_ids": ["victoria.blackwood@outlook.com", "v.blackwood@techconsult.co.uk"],
        "phones": ["+44 7911 123456"],
        "current_location": "Cambridge, United Kingdom",
        "linkedin_url": "https://linkedin.com/in/victoriablackwood",
        "github_url": "https://github.com/vblackwood",
        "total_experience_years": 15,
        "key_skills_contains": ["Java", "Python", "Go", "Kubernetes", "Kafka", "AWS", "Azure"],
        "work_experience_count": 5,
        "education_count": 2,
        "certifications_contains": ["CKA", "TOGAF"],
        "gender": None,
    },
}


def _norm_cert(item) -> str:
    """Normalise a certification entry to a plain string."""
    if isinstance(item, dict):
        return str(item.get("name") or item.get("title") or item.get("certification") or "").strip()
    return str(item).strip()


def _norm_phone_digits(phone: str) -> str:
    """Strip everything except digits for fuzzy phone comparison."""
    return re.sub(r"\D", "", str(phone))


def _check_field(extracted: dict, field: str, expected_value, tolerance_mode: str = "exact") -> dict:
    """
    Compare a single extracted field against the expected value.

    Returns dict with: field, expected, extracted, match, notes
    """
    # ── Special case: total_experience_years → parse from total_experience string ──
    if field == "total_experience_years":
        raw_exp = str(extracted.get("total_experience") or "")
        # Parse leading number, e.g. "7 years 3 months" → 7,  "0" → 0
        m = re.match(r"(\d+(?:\.\d+)?)", raw_exp.strip())
        actual = float(m.group(1)) if m else None
        result = {
            "field": field,
            "expected": str(expected_value),
            "extracted": str(actual) if actual is not None else "null",
            "match": False,
            "notes": "",
        }
        if actual is None:
            if expected_value == 0:
                result["match"] = True  # "0" experience correctly returned empty
            else:
                result["notes"] = f"Could not parse years from: {raw_exp!r}"
        else:
            result["match"] = abs(actual - float(expected_value)) <= 1
        return result

    actual = extracted.get(field)

    # Normalise certifications: list of dicts → list of strings before comparison
    if field == "certifications" and isinstance(actual, list):
        actual = [_norm_cert(x) for x in actual]
        actual = [x for x in actual if x]  # drop empty strings

    result = {
        "field": field,
        "expected": str(expected_value),
        "extracted": str(actual) if actual is not None else "null",
        "match": False,
        "notes": "",
    }

    if expected_value is None:
        result["match"] = actual is None or actual == "" or actual == [] or actual == "null"
        if result["match"]:
            result["notes"] = "Correctly null/empty"
        else:
            result["notes"] = f"Expected null, got: {actual}"
        return result

    if tolerance_mode == "exact":
        if isinstance(expected_value, str):
            # For names: strip academic title prefixes (Dr., Prof.) and suffixes (PhD, Ph.D., MBA)
            _TITLE_RE = re.compile(
                r"^(?:dr\.?|prof\.?|mr\.?|ms\.?|mrs\.?|miss)\s+|\s*,?\s*(?:ph\.?d\.?|md|mba|msc|phd|dsc|jd)$",
                re.IGNORECASE,
            )
            def _strip_titles(s: str) -> str:
                return _TITLE_RE.sub("", s.strip()).strip().lower()

            result["match"] = (
                _strip_titles(str(actual)) == _strip_titles(expected_value)
                if actual else False
            )
        elif isinstance(expected_value, (int, float)):
            try:
                result["match"] = abs(float(str(actual)) - float(expected_value)) <= 1 if actual is not None else False
            except (ValueError, TypeError):
                result["match"] = False
        elif isinstance(expected_value, list):
            actual_list = actual if isinstance(actual, list) else []
            # Phones: fuzzy digit-only comparison
            if field == "phones":
                actual_digits = {_norm_phone_digits(x) for x in actual_list}
                expected_digits = {_norm_phone_digits(x) for x in expected_value}
                # A phone matches if last 9+ digits agree (handles country-code differences)
                def _phones_match(exp_d, act_set):
                    suffix = exp_d[-9:] if len(exp_d) >= 9 else exp_d
                    return any(a.endswith(suffix) for a in act_set)
                missing = [e for e in expected_value if not _phones_match(_norm_phone_digits(e), actual_digits)]
                result["match"] = len(missing) == 0
                if missing:
                    result["notes"] = f"Missing: {missing}"
            else:
                actual_lower = {str(x).lower().strip() for x in actual_list}
                expected_lower = {str(x).lower().strip() for x in expected_value}
                result["match"] = actual_lower == expected_lower
                if not result["match"]:
                    missing = expected_lower - actual_lower
                    extra = actual_lower - expected_lower
                    notes = []
                    if missing:
                        notes.append(f"Missing: {missing}")
                    if extra:
                        notes.append(f"Extra: {extra}")
                    result["notes"] = "; ".join(notes)

    elif tolerance_mode == "contains":
        if isinstance(expected_value, list):
            actual_list = actual if isinstance(actual, list) else []
            actual_lower = {str(x).lower().strip() for x in actual_list}
            missing = []
            for exp in expected_value:
                found = any(exp.lower() in a for a in actual_lower)
                if not found:
                    missing.append(exp)
            result["match"] = len(missing) == 0
            if missing:
                result["notes"] = f"Missing: {missing}"
        elif isinstance(expected_value, str):
            result["match"] = expected_value.lower() in str(actual or "").lower()

    elif tolerance_mode == "count":
        actual_list = actual if isinstance(actual, list) else []
        result["match"] = len(actual_list) >= expected_value
        result["extracted"] = str(len(actual_list))
        if not result["match"]:
            result["notes"] = f"Expected {expected_value}, got {len(actual_list)}"

    return result


def evaluate_resume(resume_filename: str, extracted_data: dict, expected: dict) -> dict:
    """Evaluate a single resume's extraction against expected values."""
    results = []

    # Direct field comparisons
    for field in ["candidate_full_name", "current_location", "linkedin_url", "github_url", "gender"]:
        if field in expected:
            results.append(_check_field(extracted_data, field, expected[field], "exact"))

    # Email comparison
    if "email_ids" in expected:
        results.append(_check_field(extracted_data, "email_ids", expected["email_ids"], "exact"))

    # Phone comparison
    if "phones" in expected:
        results.append(_check_field(extracted_data, "phones", expected["phones"], "exact"))

    # Experience years
    if "total_experience_years" in expected:
        results.append(_check_field(extracted_data, "total_experience_years", expected["total_experience_years"], "exact"))

    # Skills (contains mode)
    if "key_skills_contains" in expected:
        results.append(_check_field(extracted_data, "key_skills", expected["key_skills_contains"], "contains"))
    elif "key_skills" in expected:
        results.append(_check_field(extracted_data, "key_skills", expected["key_skills"], "exact"))

    # Work experience count
    if "work_experience_count" in expected:
        results.append(_check_field(extracted_data, "work_experience", expected["work_experience_count"], "count"))

    # Education count
    if "education_count" in expected:
        results.append(_check_field(extracted_data, "education", expected["education_count"], "count"))

    # Certifications — "contains" so extracting extra certs doesn't fail
    if "certifications_contains" in expected:
        results.append(_check_field(extracted_data, "certifications", expected["certifications_contains"], "contains"))
    elif "certifications" in expected:
        results.append(_check_field(extracted_data, "certifications", expected["certifications"], "contains"))

    # Languages — "contains" so extra languages don't penalise
    if "languages_known" in expected:
        results.append(_check_field(extracted_data, "languages_known", expected["languages_known"], "contains"))

    # GitHub contains
    if "github_url_contains" in expected:
        results.append(_check_field(extracted_data, "github_url", expected["github_url_contains"], "contains"))

    # Calculate accuracy
    total = len(results)
    passed = sum(1 for r in results if r["match"])

    return {
        "resume": resume_filename,
        "total_fields": total,
        "passed": passed,
        "accuracy": passed / total if total > 0 else 0,
        "details": results,
    }


def _export_to_excel(summary: dict, output_dir: str) -> str:
    """
    Export evaluation results as a styled Excel workbook.

    Creates two sheets:
    - 'Comparison' : per-resume, per-field extracted vs expected with match status
    - 'Summary'    : per-resume accuracy + overall accuracy row
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.worksheet import Worksheet as _OxlWorksheet
    except ImportError:
        print("openpyxl not installed — skipping Excel export")
        return ""

    wb = Workbook()

    # ── Sheet 1: Comparison ──────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "Comparison"

    headers = ["Resume", "Field", "Expected", "Extracted", "Match", "Confidence", "Notes"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for result in summary.get("per_resume", []):
        resume_name = result.get("resume", "")
        # Grab per-field confidences from the extracted data snapshot
        field_confs = result.get("field_confidences", {})
        for detail in result.get("details", []):
            ws.cell(row=row_num, column=1, value=resume_name).border = thin_border
            ws.cell(row=row_num, column=2, value=detail["field"]).border = thin_border
            ws.cell(row=row_num, column=3, value=detail["expected"][:200]).border = thin_border
            ws.cell(row=row_num, column=4, value=detail["extracted"][:200]).border = thin_border

            match_cell = ws.cell(row=row_num, column=5, value="YES" if detail["match"] else "NO")
            match_cell.border = thin_border
            match_cell.fill = pass_fill if detail["match"] else fail_fill

            conf_val = field_confs.get(detail["field"])
            conf_cell = ws.cell(
                row=row_num, column=6,
                value=f"{conf_val:.2f}" if conf_val is not None else "—",
            )
            conf_cell.border = thin_border
            conf_cell.alignment = Alignment(horizontal="center")

            ws.cell(row=row_num, column=7, value=detail.get("notes", "")[:300]).border = thin_border
            row_num += 1

    # Auto-size columns (approx)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    assert isinstance(ws2, _OxlWorksheet)
    sum_headers = ["Resume", "Fields Checked", "Passed", "Accuracy"]
    for col_idx, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for idx, result in enumerate(summary.get("per_resume", []), 2):
        ws2.cell(row=idx, column=1, value=result.get("resume", "")).border = thin_border
        ws2.cell(row=idx, column=2, value=result.get("total_fields", 0)).border = thin_border
        ws2.cell(row=idx, column=3, value=result.get("passed", 0)).border = thin_border
        acc_cell = ws2.cell(
            row=idx, column=4,
            value=f"{result.get('accuracy', 0):.1%}",
        )
        acc_cell.border = thin_border
        acc_cell.alignment = Alignment(horizontal="center")

    # Overall row
    total_row = len(summary.get("per_resume", [])) + 2
    ws2.cell(row=total_row, column=1, value="OVERALL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=summary.get("total_field_checks", 0))
    ws2.cell(row=total_row, column=3, value=summary.get("total_passed", 0))
    ws2.cell(row=total_row, column=4, value=f"{summary.get('overall_accuracy', 0):.1%}")
    for col_idx in range(1, 5):
        ws2.cell(row=total_row, column=col_idx).font = Font(bold=True)
        ws2.cell(row=total_row, column=col_idx).border = thin_border

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 12
    ws2.freeze_panes = "A2"

    xlsx_path = os.path.join(output_dir, "evaluation_results.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


def run_evaluation_sheet(
    resumes_dir: Optional[str] = None,
    output_dir: Optional[str] = None,
) -> dict:
    """
    Run the full evaluation sheet across all sample resumes.

    Parses each sample resume, extracts data, and compares against expected values.
    """
    if resumes_dir is None:
        resumes_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "sample_data", "sample_resumes",
        )
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(__file__))

    parser = ResumeParser()
    extractor = ResumeExtractor()

    print(f"\n{'='*80}")
    print("EVALUATION SHEET — Extracted vs Expected Values")
    print(f"{'='*80}")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print(f"Resumes Directory: {resumes_dir}")
    print(f"{'='*80}\n")

    all_results = []

    for filename, expected in EXPECTED_DATA.items():
        filepath = os.path.join(resumes_dir, filename)
        if not os.path.exists(filepath):
            print(f"⚠️  Skipping {filename} (file not found)")
            continue

        ext = os.path.splitext(filename)[1]
        print(f"\n--- Processing: {filename} ---")

        # Parse
        parse_result = parser.parse(filepath, ext)
        if not parse_result["success"]:
            print(f"  ❌ Parse failed: {parse_result.get('error')}")
            all_results.append({
                "resume": filename,
                "total_fields": 0,
                "passed": 0,
                "accuracy": 0,
                "details": [],
                "error": parse_result.get("error"),
            })
            continue

        # Extract
        extracted = extractor.extract(parse_result["text"])

        # Evaluate
        result = evaluate_resume(filename, extracted, expected)
        # Attach per-field confidence snapshot for Excel export
        result["field_confidences"] = extracted.get("field_confidences", {})
        all_results.append(result)

        # Print result
        print(f"  Accuracy: {result['accuracy']:.0%} ({result['passed']}/{result['total_fields']})")
        for detail in result["details"]:
            status = "✅" if detail["match"] else "❌"
            print(f"    {status} {detail['field']}: extracted={detail['extracted'][:60]}")
            if detail["notes"]:
                print(f"       Notes: {detail['notes'][:80]}")

    # Summary
    total_fields = sum(r["total_fields"] for r in all_results)
    total_passed = sum(r["passed"] for r in all_results)
    overall_accuracy = total_passed / total_fields if total_fields > 0 else 0

    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    print(f"  Resumes evaluated: {len(all_results)}")
    print(f"  Total field checks: {total_fields}")
    print(f"  Total passed: {total_passed}")
    print(f"  Overall accuracy: {overall_accuracy:.1%}")
    print(f"{'='*80}\n")

    summary = {
        "timestamp": datetime.now().isoformat(),
        "resumes_evaluated": len(all_results),
        "total_field_checks": total_fields,
        "total_passed": total_passed,
        "overall_accuracy": overall_accuracy,
        "per_resume": all_results,
    }

    # Export JSON
    json_path = os.path.join(output_dir, "evaluation_results.json")
    with open(json_path, "w") as f:
        json.dump(summary, f, indent=2, default=str)
    print(f"JSON results: {json_path}")

    # Export CSV
    csv_path = os.path.join(output_dir, "evaluation_results.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Resume", "Field", "Expected", "Extracted", "Match", "Notes"])
        for result in all_results:
            for detail in result.get("details", []):
                writer.writerow([
                    result["resume"],
                    detail["field"],
                    detail["expected"],
                    detail["extracted"],
                    "YES" if detail["match"] else "NO",
                    detail.get("notes", ""),
                ])
        # Summary row
        writer.writerow([])
        writer.writerow(["OVERALL", "", "", "", f"{overall_accuracy:.1%}", f"{total_passed}/{total_fields}"])
    print(f"CSV results:  {csv_path}")

    # Export Excel (.xlsx)
    xlsx_path = _export_to_excel(summary, output_dir)
    if xlsx_path:
        print(f"Excel results: {xlsx_path}")

    return summary


if __name__ == "__main__":
    run_evaluation_sheet()
