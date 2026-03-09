"""
Excel storage service for resume data.

Stores extracted resume data into an Excel workbook (.xlsx) using openpyxl.
Each processed resume creates or updates a row keyed by candidate_id.
Supports duplicate handling with version tracking.
"""
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from django.conf import settings

logger = logging.getLogger("api")

# Column definitions for the Excel sheet
EXCEL_COLUMNS = [
    "CandidateID",
    "CandidateFullName",
    "EmailIDs",
    "Phones",
    "Gender",
    "CurrentLocation",
    "GeoDetails",
    "TotalExperience",
    "WorkExperience",
    "KeySkills",
    "Education",
    "Certifications",
    "LinkedInURL",
    "GitHubURL",
    "PortfolioURL",
    "ProfessionalSummary",
    "Projects",
    "ExtractionConfidence",
    "FieldConfidences",
    "ResumeSource",
    "OriginalFileName",
    "ProcessedTimestamp",
    "Version",
    "OriginalFileLink",
]

# Column definitions for the ProcessingLogs sheet
LOG_COLUMNS = [
    "Timestamp",
    "RecordID",
    "CandidateID",
    "OriginalFileName",
    "Level",
    "Step",
    "Message",
    "Details",
]

# Schema definition exposed via API
EXCEL_SCHEMA = {
    "storage_type": "Excel (.xlsx)",
    "file_location": "Configured via EXCEL_OUTPUT_PATH setting",
    "description": "Stores parsed and extracted resume information in an Excel workbook",
    "sheets": [
        {
            "name": "Resume Data",
            "description": "One row per candidate — primary extracted resume information",
        },
        {
            "name": "ProcessingLogs",
            "description": "Append-only audit log for every processing step across all resumes",
        },
    ],
    "columns": [
        {"name": "CandidateID", "type": "Text", "description": "Unique candidate identifier (email+phone hash)"},
        {"name": "CandidateFullName", "type": "Text", "description": "Full name of the candidate"},
        {"name": "EmailIDs", "type": "Text (JSON)", "description": "JSON array of email addresses"},
        {"name": "Phones", "type": "Text (JSON)", "description": "JSON array of phone numbers"},
        {"name": "Gender", "type": "Text", "description": "Gender (only if explicitly present in resume)"},
        {"name": "CurrentLocation", "type": "Text", "description": "Current city/state/country"},
        {"name": "GeoDetails", "type": "Text (JSON)", "description": "Normalized city/state/country object"},
        {"name": "TotalExperience", "type": "Text", "description": "Total years/months of experience"},
        {"name": "WorkExperience", "type": "Text (JSON)", "description": "JSON array of work experience entries"},
        {"name": "KeySkills", "type": "Text (JSON)", "description": "JSON array of skills"},
        {"name": "Education", "type": "Text (JSON)", "description": "JSON array of education entries"},
        {"name": "Certifications", "type": "Text (JSON)", "description": "JSON array of certifications"},
        {"name": "LinkedInURL", "type": "Text", "description": "LinkedIn profile URL"},
        {"name": "GitHubURL", "type": "Text", "description": "GitHub profile URL"},
        {"name": "PortfolioURL", "type": "Text", "description": "Portfolio website URL"},
        {"name": "ProfessionalSummary", "type": "Text", "description": "Generated professional summary (3-5 lines)"},
        {"name": "Projects", "type": "Text (JSON)", "description": "JSON array of projects"},
        {"name": "ExtractionConfidence", "type": "Decimal", "description": "Overall extraction confidence (0.0-1.0)"},
        {"name": "FieldConfidences", "type": "Text (JSON)", "description": "Per-field confidence scores"},
        {"name": "ResumeSource", "type": "Text", "description": "Source: 'upload' or 'email'"},
        {"name": "OriginalFileName", "type": "Text", "description": "Original uploaded file name"},
        {"name": "ProcessedTimestamp", "type": "DateTime", "description": "When the resume was processed"},
        {"name": "Version", "type": "Integer", "description": "Record version (incremented on resubmission)"},
        {"name": "OriginalFileLink", "type": "Text", "description": "Link/reference to original file on disk"},
    ],
    "log_columns": [
        {"name": "Timestamp", "type": "DateTime", "description": "UTC timestamp of the log entry"},
        {"name": "RecordID", "type": "Text", "description": "UUID of the associated ResumeRecord"},
        {"name": "CandidateID", "type": "Text", "description": "Unique candidate identifier"},
        {"name": "OriginalFileName", "type": "Text", "description": "Original uploaded file name"},
        {"name": "Level", "type": "Text", "description": "Log level: info / warning / error / debug"},
        {"name": "Step", "type": "Text", "description": "Processing step name"},
        {"name": "Message", "type": "Text", "description": "Human-readable log message"},
        {"name": "Details", "type": "Text (JSON)", "description": "Additional structured details"},
    ],
}


class ExcelStorageClient:
    """
    Client for storing resume data in an Excel workbook.

    Uses openpyxl to read/write .xlsx files. Handles:
    - Creating the workbook and header row if it doesn't exist
    - Inserting new rows for new candidates
    - Updating existing rows (upsert by candidate_id)
    - Version tracking for duplicate resumes
    """

    def __init__(self):
        self.excel_path = getattr(settings, "EXCEL_OUTPUT_PATH", None)
        if not self.excel_path:
            self.excel_path = os.path.join(settings.BASE_DIR, "outputs", "resume_data.xlsx")
        # Ensure directory exists
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)

    @property
    def is_configured(self) -> bool:
        """Excel storage is always available (local file)."""
        return True

    def _get_or_create_workbook(self):
        """Open existing workbook or create a new one with headers."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if os.path.exists(self.excel_path):
            try:
                wb = load_workbook(self.excel_path)
                ws = wb.active
                # Ensure the ProcessingLogs sheet exists in older workbooks
                if "ProcessingLogs" not in wb.sheetnames:
                    self._create_logs_sheet(wb)
                    wb.save(self.excel_path)
                return wb, ws
            except Exception as e:
                logger.warning(f"Could not open existing Excel file, creating new: {e}")

        # Create new workbook with styled header
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume Data"

        # Style the header row
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set reasonable column widths
        widths = {
            "A": 18, "B": 25, "C": 30, "D": 25, "E": 10, "F": 25,
            "G": 20, "H": 18, "I": 40, "J": 40, "K": 30, "L": 30,
            "M": 35, "N": 35, "O": 35, "P": 50, "Q": 30, "R": 15,
            "S": 20, "T": 12, "U": 30, "V": 22, "W": 10, "X": 40,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add the ProcessingLogs sheet
        self._create_logs_sheet(wb)

        wb.save(self.excel_path)
        logger.info(f"Created new Excel workbook: {self.excel_path}")
        return wb, ws

    def _create_logs_sheet(self, wb):
        """Add a styled ProcessingLogs sheet to the workbook."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws_logs = wb.create_sheet(title="ProcessingLogs")

        log_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        log_header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
        log_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(LOG_COLUMNS, start=1):
            cell = ws_logs.cell(row=1, column=col_idx, value=col_name)
            cell.font = log_header_font
            cell.fill = log_header_fill
            cell.alignment = log_header_alignment
            cell.border = thin_border

        log_widths = {"A": 22, "B": 38, "C": 22, "D": 30, "E": 10, "F": 18, "G": 60, "H": 40}
        for col_letter, width in log_widths.items():
            ws_logs.column_dimensions[col_letter].width = width

        ws_logs.freeze_panes = "A2"
        return ws_logs

    def append_log(
        self,
        level: str,
        step: str,
        message: str,
        record_id: str = "",
        candidate_id: str = "",
        original_file_name: str = "",
        details: Optional[dict] = None,
    ) -> None:
        """
        Append a single log entry to the ProcessingLogs sheet.

        Args:
            level: Log level (info / warning / error / debug)
            step: Processing step name
            message: Human-readable log message
            record_id: UUID of the associated ResumeRecord (optional)
            candidate_id: Unique candidate identifier (optional)
            original_file_name: Original uploaded file name (optional)
            details: Additional structured details (optional)
        """
        try:
            wb, _ws = self._get_or_create_workbook()
            ws_logs = wb["ProcessingLogs"]
            row = [
                datetime.utcnow().isoformat(),
                str(record_id),
                str(candidate_id),
                str(original_file_name),
                str(level),
                str(step),
                str(message),
                json.dumps(details or {}, ensure_ascii=False),
            ]
            ws_logs.append(row)
            wb.save(self.excel_path)
        except Exception as e:
            logger.error(f"Excel append_log error: {e}")

    def _find_row_by_candidate_id(self, ws, candidate_id: str) -> Optional[int]:
        """Find the row number for a given candidate_id (column A)."""
        if not candidate_id:
            return None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == candidate_id:
                return row
        return None

    def _map_to_row(self, data: dict) -> list:
        """Map extracted resume data to a row of cell values matching EXCEL_COLUMNS."""
        def _json_dump(value):
            if isinstance(value, (list, dict)):
                return json.dumps(value, ensure_ascii=False)
            return str(value) if value else ""

        return [
            data.get("candidate_id", ""),
            data.get("candidate_full_name", ""),
            _json_dump(data.get("email_ids", [])),
            _json_dump(data.get("phones", [])),
            data.get("gender", ""),
            data.get("current_location", ""),
            _json_dump(data.get("geo_details", {})),
            data.get("total_experience", ""),
            _json_dump(data.get("work_experience", [])),
            _json_dump(data.get("key_skills", [])),
            _json_dump(data.get("education", [])),
            _json_dump(data.get("certifications", [])),
            data.get("linkedin_url", ""),
            data.get("github_url", ""),
            data.get("portfolio_url", ""),
            data.get("professional_summary", ""),
            _json_dump(data.get("projects", [])),
            data.get("extraction_confidence", 0.0),
            _json_dump(data.get("field_confidences", {})),
            data.get("resume_source", "upload"),
            data.get("original_file_name", ""),
            datetime.utcnow().isoformat(),
            data.get("version", 1),
            data.get("file_path", ""),
        ]

    def create_record(self, data: dict) -> dict:
        """
        Create a new row in the Excel workbook.

        Args:
            data: Extracted resume data dict

        Returns:
            dict with status and row info
        """
        try:
            wb, ws = self._get_or_create_workbook()
            row_values = self._map_to_row(data)
            ws.append(row_values)
            wb.save(self.excel_path)

            row_number = ws.max_row
            candidate_id = data.get("candidate_id", "")
            logger.info(f"Excel record created at row {row_number}: {candidate_id}")

            return {
                "id": candidate_id,
                "row": row_number,
                "status": "created",
                "action": "created",
                "version": data.get("version", 1),
            }
        except Exception as e:
            logger.error(f"Excel create error: {e}")
            return {"id": None, "status": "error", "error": str(e)}

    def update_record(self, row_number: int, data: dict) -> dict:
        """
        Update an existing row in the Excel workbook.

        Args:
            row_number: The Excel row to update (1-based, >=2)
            data: Updated resume data

        Returns:
            dict with status
        """
        try:
            wb, ws = self._get_or_create_workbook()
            row_values = self._map_to_row(data)

            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_number, column=col_idx, value=value)

            wb.save(self.excel_path)
            candidate_id = data.get("candidate_id", "")
            logger.info(f"Excel record updated at row {row_number}: {candidate_id}")

            return {
                "id": candidate_id,
                "row": row_number,
                "status": "updated",
                "action": "updated",
                "version": data.get("version", 1),
            }
        except Exception as e:
            logger.error(f"Excel update error: {e}")
            return {"id": None, "status": "error", "error": str(e)}

    def find_by_candidate_id(self, candidate_id: str) -> Optional[dict]:
        """
        Find an existing record by candidate_id.

        Args:
            candidate_id: Unique candidate identifier

        Returns:
            dict with row data if found, None otherwise
        """
        if not candidate_id:
            return None

        try:
            wb, ws = self._get_or_create_workbook()
            row_num = self._find_row_by_candidate_id(ws, candidate_id)

            if row_num is None:
                return None

            # Build dict from row
            record = {}
            for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
                record[col_name] = ws.cell(row=row_num, column=col_idx).value

            record["_row_number"] = row_num
            logger.info(f"Found existing Excel record for candidate: {candidate_id}")
            return record

        except Exception as e:
            logger.error(f"Excel find error: {e}")
            return None

    def upsert_record(self, data: dict) -> dict:
        """
        Insert or update a record based on candidate_id.
        Handles duplicate detection and version tracking.

        Opens the workbook exactly ONCE per call so that the find and
        write operations share the same in-memory state, preventing
        stale-read races that could otherwise cause duplicate rows.

        Args:
            data: Extracted resume data

        Returns:
            dict with upsert result
        """
        try:
            wb, ws = self._get_or_create_workbook()
            candidate_id = data.get("candidate_id", "")

            if candidate_id:
                row_num = self._find_row_by_candidate_id(ws, candidate_id)
                if row_num:
                    # --- Update existing row ---
                    version_col = EXCEL_COLUMNS.index("Version") + 1
                    current_version = ws.cell(row=row_num, column=version_col).value or 1
                    try:
                        current_version = int(current_version)
                    except (ValueError, TypeError):
                        current_version = 1
                    data["version"] = current_version + 1
                    row_values = self._map_to_row(data)
                    for col_idx, value in enumerate(row_values, start=1):
                        ws.cell(row=row_num, column=col_idx, value=value)
                    wb.save(self.excel_path)
                    logger.info(
                        f"Excel record updated at row {row_num}: {candidate_id} "
                        f"(version {data['version']})"
                    )
                    return {
                        "id": candidate_id,
                        "row": row_num,
                        "status": "updated",
                        "action": "updated",
                        "version": data["version"],
                    }

            # --- Insert new row ---
            data.setdefault("version", 1)
            row_values = self._map_to_row(data)
            ws.append(row_values)
            wb.save(self.excel_path)
            row_number = ws.max_row
            logger.info(f"Excel record created at row {row_number}: {candidate_id}")
            return {
                "id": candidate_id,
                "row": row_number,
                "status": "created",
                "action": "created",
                "version": data.get("version", 1),
            }

        except Exception as e:
            logger.error(f"Excel upsert error: {e}")
            return {"id": None, "status": "error", "error": str(e)}

    def get_all_records(self) -> list:
        """
        Read all records from the Excel workbook.

        Returns:
            list of dicts
        """
        try:
            wb, ws = self._get_or_create_workbook()
            records = []

            for row in range(2, ws.max_row + 1):
                # Skip empty rows
                if ws.cell(row=row, column=1).value is None:
                    continue
                record = {}
                for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
                    record[col_name] = ws.cell(row=row, column=col_idx).value
                records.append(record)

            return records
        except Exception as e:
            logger.error(f"Excel read all error: {e}")
            return []

    def get_record_count(self) -> int:
        """Return the number of records in the Excel file."""
        try:
            wb, ws = self._get_or_create_workbook()
            count = 0
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value is not None:
                    count += 1
            return count
        except Exception:
            return 0
